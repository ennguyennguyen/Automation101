# %% -- IMPORT LIBRARIES
import openpyxl

# %% -- READ DATA FROM EXCEL FILE
path = "D:\\00_MASTER OF COMPUTER SCIENCE_MUM\\00_Projects\\12_Automation Testing_Basic\\DataDriven_Input.xlsx"

workbook = openpyxl.load_workbook(path)
sheet = workbook.get_sheet_by_name("Sheet1") # or use workbook.active to get the current worksheet

rows = sheet.max_row        # max_row will return all row
cols = sheet.max_column     # max_column will return all column

for row in range(1, rows+1):
    for col in range(1, cols+1):
        print(sheet.cell(row = row, column = col).value, end = "    ")      # use sheet.cell(row, col).value to retrieve cell value
    print()

# %% -- WRITE DATA TO EXCEL FILE
path = "D:\\00_MASTER OF COMPUTER SCIENCE_MUM\\00_Projects\\12_Automation Testing_Basic\\DataDriven_Output.xlsx"

workbook = openpyxl.load_workbook(path)
sheet = workbook.active

for r in range(1, 10):
    for c in range(1, 4):
        sheet.cell(row = r, column = c).value = "hello"

workbook.save(path)     # save to workbook


