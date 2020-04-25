import openpyxl

def countRow(file, sName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sName)
    return(sheet.max_row)


def countCol(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    return(sheet.max_column)

def readFile(file, sheetName, numRow, numCol):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    return sheet.cell(row = numRow, column = numCol).value

def writeFile(file, sheetName, numRow, numCol, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    sheet.cell(row = numRow, column = numCol).value = data
    workbook.save(file)